from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Cargar el archivo Excel
archivo_excel = 'PEDIDITO_DB.xlsx'
hoja_nombre = 'BASE_PEDIDITO'

wb = load_workbook(archivo_excel)
hoja = wb[hoja_nombre]

# Crear una nueva fila con tus datos
nueva_fila = ["31/08/2023","No especificado","Melissa","5624838493","jaula buba","6b 201","garrafon equly",32.00,22.00,10.00,"No especificado",""]  # Ajusta los valores según tus columnas


# Insertar la nueva fila al principio de la tabla
hoja.insert_rows(2, amount=1)
for col_num, value in enumerate(nueva_fila, start=1):
    hoja.cell(row=2, column=col_num, value=value)

# Guardar el archivo Excel actualizado
wb.save(archivo_excel)
